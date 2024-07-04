import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import *
import os

def read_and_process_sheets(file_path):
    """读取Excel数据并处理每个sheet中的指定列"""
    xls = pd.ExcelFile(file_path)
    processed_dfs = {}
    
    for sheet_name in xls.sheet_names:
        # 读取单个sheet的数据
        df = xls.parse(sheet_name)
        
        # 处理数据，删除指定的列
        df = process_data(df)
        
        # 存储处理后的DataFrame
        processed_dfs[sheet_name] = df
    
    return processed_dfs

def read_and_process_sheets_up(file_path):
     xls = pd.ExcelFile(file_path)
     df = xls.parse('分区县')
     df = process_data_up(df)
     return df

def process_data(data):
    """处理数据，删除不需要的列，排序等"""
    # 删除指定的列
    data.drop(columns=['已上传辅助证件合规数量', '上传合规率'], inplace=True)
    # 将'上传率'列转换为百分比形式
    if '上传率' in data.columns:  # 确保'上传率'列存在
        data['上传率'] = data['上传率'].apply(lambda x: f"{x*100:.2f}%")  # 转换为百分比并格式化为两位小数点后加%
    return data

def process_data_up(data):
    """处理数据，删除不需要的列，补充缺失区县数据，并确保全通路在最后"""
    required_counties = ['湖里', '思明', '海沧', '集美', '同安', '营业外包', '翔安']
    
    # 确保全通路记录在最后一行
    if '全通路' in data['区县'].values:
        all_path_row = data[data['区县'] == '全通路']
        data = data[data['区县'] != '全通路']
        if not data.empty and all_path_row.empty:
            all_path_row = all_path_row.iloc[[0]]  # 处理只有一行'全通路'的情况
    
    # 补充缺失的区县数据
    existing_counties = set(data['区县'])
    missing_counties = set(required_counties) - existing_counties
    for county in missing_counties:
        new_row = pd.DataFrame({
            '区县': [county],
            '应上传辅助证件数量': [0]
        })
        data = pd.concat([data, new_row], ignore_index=True)

    # 先将百分比字符串转换为浮点数
    data.loc[:, '上传率'] = data['上传率'].str.rstrip('%').astype(float) / 100

    # 按上传率降序排列（假设上传率列存在且需要排序）
    if '上传率' in data.columns:
        data_sorted = data.sort_values(by='上传率', ascending=False)
        data_sorted = data_sorted[~data_sorted['区县'].isin(missing_counties)]  # 排除刚添加的0值行
        data = pd.concat([data_sorted, data[data['区县'].isin(missing_counties)]], ignore_index=True)  # 保证0值行在正确位置

    # 排序完成后，再转换回百分比格式
    data['上传率'] = data['上传率'].apply(lambda x: f"{x * 100:.2f}%")

     # 确保全通路行在最后
    if not all_path_row.empty:
        data = pd.concat([data, all_path_row], ignore_index=True)

    return data

def save_processed_sheets(processed_dfs, output_path):
    """保存处理后的所有sheet到一个新的Excel文件"""
    with pd.ExcelWriter(output_path) as writer:
        for sheet_name, df in processed_dfs.items():
            #df = df.loc[:, ~df.columns.str.startswith('Unnamed: ')]
            #df.rename(columns={'Unnamed: 0':''})
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    size_by_content(output_path)

def save_processed_sheets_up(processed_dfs_day, output_path):

    """保存处理后到一个新的Excel文件"""
    processed_dfs_day.to_excel(output_path, index=False)
    size_by_content(output_path)
    

def save_file(input_file_path,output_file_path):
    # 读取并处理所有sheet
    processed_dfs = read_and_process_sheets(input_file_path)
    
    # 保存处理后的数据到新文件
    save_processed_sheets(processed_dfs, output_file_path)
    print("处理并保存完成。")

def save_file_up(input_file_path,output_file_path_up):
    processed_dfs = read_and_process_sheets_up(input_file_path)
    save_processed_sheets_up(processed_dfs,output_file_path_up)
    print("处理并保存完成。")

def size_by_content(file_path,file = None):
    # 加载Excel文件
    if file:
        workbook = file
    else:
        workbook = load_workbook(file_path)
    #居中格式
    align=Alignment(horizontal='center',vertical='center')

    # 设置边框显示
    border = Border(top=Side(border_style='thin', color='FF000000'),
           right=Side(border_style='thin', color='FF000000'),
           bottom=Side(border_style='thin', color='FF000000'),
           left=Side(border_style='thin', color='FF000000'))

    # 定义字体样式
    font_style = Font(name='微软雅黑', size=12)

    # 获取每个Sheet并调整列宽
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for column_cells in sheet.columns:
            length=0
            for cell in column_cells:
                length = max(len(str(cell.value)),length)
                cell.alignment = align
                cell.font = font_style
                cell.border = border
                if cell.value=='Unnamed: 0':
                    cell.value=''
            #length = max(len(str(cell.value)) for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = length*2+5
        if file:
            return workbook
    workbook.save(file_path)
    workbook.close()
    
def save_file_up_concat(output_file_path_day_up,output_file_path_month_up,output_file_path_up,title_day,title_month,formatted_two_days_ago):
    #formatted_two_days_ago='6.29'
    #居中格式
    align=Alignment(horizontal='center',vertical='center')
    # 设置边框显示
    border = Border(top=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='thin', color='FF000000'),
                    left=Side(border_style='thin', color='FF000000'))
    # 定义字体样式
    title_font = Font(name='微软雅黑', size=11, bold=True)
    # 加载现有的Excel文件
    wb1 = load_workbook(output_file_path_day_up)
    wb2 = load_workbook(output_file_path_month_up)
    # 选择要合并的Sheet
    sheet1 = wb1.worksheets[0]
    sheet2 = wb2.worksheets[0]

    # 检查文件是否存在
    if os.path.exists(output_file_path_up):
        # 如果文件存在，则加载工作簿
        wb = load_workbook(filename=output_file_path_up)
    else:
        # 创建一个新的Excel工作簿
        wb = Workbook()
    # 删除默认的'Sheet'工作表
    if 'Sheet' in wb.sheetnames:
        default_sheet = wb['Sheet']
        wb.remove(default_sheet)
    if formatted_two_days_ago in wb.sheetnames:
        formatted_two_days_ago_sheet = wb[formatted_two_days_ago]
        wb.remove(formatted_two_days_ago_sheet)
    # 创建新的工作表
    ws = wb.create_sheet(title=formatted_two_days_ago)
    if formatted_two_days_ago in wb.sheetnames:
        formatted_two_days_ago_sheet = wb[formatted_two_days_ago]
        wb.remove(formatted_two_days_ago_sheet)
    wb._sheets.insert(0, ws)
    # 复制第一个Sheet的数据到新Sheet
    for row in sheet1.iter_rows():
        row_data = [cell.value for cell in row]
        ws.append(row_data)
   
    # 复制第二个Sheet的数据到新Sheet
    for row in sheet2.iter_rows():
        row_data = [cell.value for cell in row]
        ws.append(row_data)
    print(wb.sheetnames)
    wb = size_by_content(output_file_path_up,wb)
    ws = wb[formatted_two_days_ago]
    # 给表格添加标题
    if title_day :
        ws.insert_rows(1)  # 在第一行前插入新行作为标题行
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet1.max_column)  # 合并第一行的所有单元格
        ws.cell(row=1, column=1).value = title_day  # 设置合并后的单元格值
        ws.cell(row=1, column=1).alignment = align
        ws.cell(row=1, column=1).font = title_font  # 设置字体
        ws.cell(row=1, column=1).border = border
    if title_month :
        ws.insert_rows(2+sheet1.max_row)  # 在第一行前插入新行作为标题行
        ws.merge_cells(start_row=2+sheet1.max_row, start_column=1, end_row=2+sheet1.max_row, end_column=sheet2.max_column)  # 合并第一行的所有单元格
        ws.cell(row=2+sheet1.max_row, column=1).value = title_month  # 设置合并后的单元格值
        ws.cell(row=2+sheet1.max_row, column=1).alignment = align
        ws.cell(row=2+sheet1.max_row, column=1).font = title_font  # 设置字体
        ws.cell(row=2+sheet1.max_row, column=1).border = border
    # 保存修改后的Excel文件
    # 遍历工作簿中的所有工作表
    to_remove = []
    for sheet in wb:
        # 检查工作表名是否以"Recovered_Sheet"开头
        if sheet.title.startswith('Recovered_Sheet'):
            to_remove.append(sheet)
    # 删除匹配的工作表
    for sheet in to_remove:
        wb.remove(sheet)
    print(wb.sheetnames)
    wb.save(filename=output_file_path_up)
    wb.close()
    wb1.close()
    wb2.close()
def main():
    # 获取当前日期
    today = datetime.today()
    # 计算前两天的日期
    two_days_ago = today - relativedelta(days=4)
    current_month = str(two_days_ago.month)  # 格式化为6形式
    current_day = two_days_ago.strftime('%d')
    formatted_two_days_ago = f'{current_month}.{current_day}'  # 格式化为6.27形式
    formatted_two_days_ago2 = f'{current_month}.01-{current_day}'
    #文件根目录
    root_directory = r'F:/IPA_project/data/'
    #新入网辅助证件统计表按日
    input_file_path_day = f'{root_directory}新入网辅助证件统计表{formatted_two_days_ago}.xlsx'
    output_file_path_day = f'{root_directory}新入网辅助证件统计表按日.xlsx'
    save_file(input_file_path_day,output_file_path_day)


    #新入网辅助证件统计表按月
    input_file_path_month = f'{root_directory}新入网辅助证件统计表{formatted_two_days_ago2}.xlsx'
    output_file_path_month = f'{root_directory}新入网辅助证件统计表按月.xlsx'
    save_file(input_file_path_month,output_file_path_month)

    #X月辅助证件上传率日通报
    output_file_path_up = f'{root_directory}{current_month}月辅助证件上传率日通报.xlsx'
    output_file_path_day_up=f'{root_directory}{current_month}月辅助证件上传率日通报_day.xlsx'
    output_file_path_month_up=f'{root_directory}{current_month}月辅助证件上传率日通报_month.xlsx'
    title_day = f'{current_month}月{current_day}日辅助上传情况'
    title_month = f'{current_month}.01-6.{current_day}辅助证件上传情况（180天口径）'
    save_file_up(output_file_path_day,output_file_path_day_up)
    save_file_up(output_file_path_month,output_file_path_month_up)
    save_file_up_concat(output_file_path_day_up,output_file_path_month_up,output_file_path_up,title_day,title_month,formatted_two_days_ago)

if __name__ == "__main__":
    main()