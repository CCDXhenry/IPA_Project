from openpyxl import load_workbook

# 加载Excel文件
workbook = load_workbook(filename='240726降档报表_自助取数.xlsx')
sheet = workbook.active

# 本月降档累计
monthly_degraded = int(sheet['F10'].value)

# 环比增长
delta_degraded = int(sheet['G10'].value)

# 环比增幅
previous_month_total = int(sheet['N10'].value)
growth_rate = (delta_degraded / previous_month_total) * 100 if previous_month_total != 0 else 0

# 提取区域数据
regions_data = []
for row in sheet.iter_rows(min_row=4, max_row=9, min_col=3, max_col=14):
    region, ring_growth, prev_total, degradation = row[0].value, float(row[4].value), float(row[11].value), float(row[10].value)
    regions_data.append({
        'region': region,
        'ring_growth': ring_growth,
        'prev_total': prev_total,
        'degradation': degradation
    })

# 计算环比变化百分比
for region in regions_data:
    region['ring_ratio_change'] = (region['ring_growth'] / region['prev_total']) * 100 if region['prev_total'] != 0 else 0

# 找到环比变化最好的区域
best_region = min(regions_data, key=lambda x: x['ring_ratio_change'])

# 找到环比变化最差的两个区域
worst_regions = sorted(regions_data, key=lambda x: x['ring_ratio_change'], reverse=True)[:2]

# 户均降档套餐费恶化
average_degradation = float(sheet['M10'].value)
# 格式化户均降档套餐费恶化的金额
average_degradation_str = f"{average_degradation:.2f}"  # 默认保留两位小数
average_degradation_str = average_degradation_str.rstrip('0').rstrip('.')  # 移除末尾的零和小数点

degradation_regions = sorted(regions_data, key=lambda x: x['degradation'], reverse=True)[:2]
'''
# 输出结果
output = (
    f"1.本月降档累计 {monthly_degraded/10000:.1f}万 笔，环比增长 {int(delta_degraded/100)*100} 笔，增幅 {growth_rate:.0f}%"
    f"{best_region['region']} 相对较好，{'、'.join([r['region'] for r in worst_regions])} 环比恶化较多。"
    f"\n户均降档套餐费恶化 {average_degradation_str} 元，{'、'.join([r['region'] for r in degradation_regions])}恶化相对较高。"
)
print(output)
'''
'''
print(f"本月降档累计 {monthly_degraded/10000:.1f}万 笔，环比增长 {int(delta_degraded/100)*100} 笔，增幅 {growth_rate:.0f}%")
print(f"{best_region['region']} 相对较好，{'、'.join([r['region'] for r in worst_regions])} 环比恶化较多。")
print(f"户均降档套餐费恶化 {average_degradation_str} 元，{'、'.join([r['region'] for r in degradation_regions])}恶化相对较高。")
'''
passage_data = []
for row in sheet.iter_rows(min_row=13, max_row=24, min_col=4, max_col=10):
    passage_data.append(row)

passage_data[0][0].value = '客服'
# 获取当前月和上个月的数据
now_month = [cell.value for cell in passage_data[9]]
old_month = [cell.value for cell in passage_data[11]]

# 计算环比变化百分比
ratio_changes1 = [(float(now_month[i]) - float(old_month[i])) if float(old_month[i]) != 0 else 0 for i in range(len(now_month))]
ratio_changes2 = [(float(now_month[i]) - float(old_month[i])) / float(old_month[i]) if float(old_month[i]) != 0 else 0 for i in range(len(now_month))]

# 找到最大环比变化百分比的索引
max_index = ratio_changes1.index(max(ratio_changes1))

# 保存i=1和i=5时的值
service_value = ratio_changes2[1]
Internet_value = ratio_changes2[5]
'''
# 输出结果
output = (
    f"1.本月降档累计 {monthly_degraded/10000:.1f}万 笔，环比增长 {int(delta_degraded/100)*100} 笔，增幅 {growth_rate:.0f}%"
    f"{best_region['region']} 相对较好，{'、'.join([r['region'] for r in worst_regions])} 环比恶化较多。"
    f"户均降档套餐费恶化 {average_degradation_str} 元，{'、'.join([r['region'] for r in degradation_regions])}恶化相对较高。"
    f"\n2、按通路来看，{passage_data[0][max_index].value}通路环比增长较多，主要系客户合约到期投诉降档，互联网通路降档量环比增幅{Internet_value*100:.0f}%，营业厅降档环比增幅{service_value*100:.0f}%，请区县主抓合约到期接续场景，加强到期客户接续营销服务。"
)
print(output)
'''


def calculate_ratio(numerator_cell, denominator_cell):
    """Calculate the ratio of two cells' values."""
    return (float(numerator_cell.value) / float(denominator_cell.value)) * 100 if denominator_cell.value != 0 else 0


def calculate_list_ratio(sheet, row1, row2, region_row):
    """Calculate the list-based ratio and find the top two regions."""
    list1 = [cell.value for cell in sheet[f'J{row1}':f'O{row1}'][0]]
    list2 = [cell.value for cell in sheet[f'J{row2}':f'O{row2}'][0]]
    regions = [cell.value for cell in sheet[f'J{region_row}':f'O{region_row}'][0]]

    # Calculate the ratio for each element in the lists
    ratios = [(float(list1[i]) - float(list2[i])) / float(list2[i]) if float(list2[i]) != 0 else 0 for i in
              range(len(list2))]

    # Find the indices of the top two ratios
    top_two_indices = sorted(range(len(ratios)), key=lambda i: ratios[i], reverse=True)[:2]

    # Get the corresponding regions
    top_two_regions = [regions[index] for index in top_two_indices]

    return ratios, top_two_regions


# High value degraded ratio
high_value_degraded_ratio = calculate_ratio(sheet['E116'], sheet['E92'])

# High value regions
high_value_ratios, high_value_top_regions = calculate_list_ratio(sheet, 116, 92, 91)

# Business self-substitution ratio
self_substitution_ratio = calculate_ratio(sheet['E183'], sheet['E177'])

# Business self-substitution regions
self_substitution_ratios, self_substitution_top_regions = calculate_list_ratio(sheet, 183, 177, 91)

# Network competition ratio
network_competition_ratio = calculate_ratio(sheet['E187'], sheet['E177'])

# Network competition regions
network_competition_ratios, network_competition_top_regions = calculate_list_ratio(sheet, 187, 177, 91)

# Group degraded ratio
group_degraded_ratio = calculate_ratio(sheet['E101'], sheet['E92'])

# Group degraded regions
group_degraded_ratios, group_degraded_top_regions = calculate_list_ratio(sheet, 101, 92, 91)
# 输出结果
output = (
    f"1.本月降档累计{monthly_degraded/10000:.1f}万笔，环比增长{int(delta_degraded/100)*100}笔，增幅{growth_rate:.0f}%"
    f"{best_region['region']} 相对较好，{'、'.join([r['region'] for r in worst_regions])}环比恶化较多。"
    f"户均降档套餐费恶化{average_degradation_str} 元，{'、'.join([r['region'] for r in degradation_regions])}恶化相对较高。"
    f"\n2、按通路来看，{passage_data[0][max_index].value}通路环比增长较多，主要系客户合约到期投诉降档，互联网通路降档量环比增幅{Internet_value*100:.0f}%，营业厅降档环比增幅{service_value*100:.0f}%，请区县主抓合约到期接续场景，加强到期客户接续营销服务。"
    f"\n3、降档归因及管控：高价值降档占比{high_value_degraded_ratio:.0f}%,{'、'.join(high_value_top_regions)}最高，"
    f"业务自替代占比{self_substitution_ratio:.1f}%，{'、'.join(self_substitution_top_regions)}相对严重，"
    f"异网抢夺{network_competition_ratio:.1f}%，{'、'.join(network_competition_top_regions)}手竟对干扰较严重。"
    f"集团降档占比 {group_degraded_ratio:.0f}%，{'、'.join(group_degraded_top_regions)}、重客集团降档较多需加强保有。")
output = (
    f"1. 本月降档累计{monthly_degraded/10000:.1f}万笔，"
    f"环比{'增长' if delta_degraded > 0 else '减少'}{abs(delta_degraded)}笔，"
    f"{'增幅' if growth_rate >= 0 else '降幅'}{abs(growth_rate):.0f}%，"
    f"{best_region['region']}相对较好，{'、'.join([r['region'] for r in worst_regions])}环比恶化较多。"
    f"户均降档套餐费恶化{average_degradation_str}元，{'、'.join([r['region'] for r in degradation_regions])}恶化相对较高。"
    f"\n2. 按通路来看，{passage_data[0][max_index].value}通路环比增长较多，"
    f"主要系客户合约到期投诉降档，互联网通路降档量环比{'增幅' if Internet_value >= 0 else '降幅'}{abs(Internet_value*100):.0f}%，"
    f"营业厅降档环比{'增幅' if service_value >= 0 else '降幅'}{abs(service_value*100):.0f}%，"
    f"请区县主抓合约到期接续场景，加强到期客户接续营销服务。"
    f"\n3. 降档归因及管控：高价值降档占比{high_value_degraded_ratio:.0f}%, "
    f"{'、'.join(high_value_top_regions)}最高，"
    f"业务自替代占比{self_substitution_ratio:.1f}%，{'、'.join(self_substitution_top_regions)}相对严重，"
    f"异网抢夺{network_competition_ratio:.1f}%，{'、'.join(network_competition_top_regions)}手竟对干扰较严重。"
    f"集团降档占比{group_degraded_ratio:.0f}%，{'、'.join(group_degraded_top_regions)}、重客集团降档较多需加强保有。"
)


print(output)
'''
# 高价值降档占比
high_value_degraded_ratio = (float(sheet['E116'].value) / float(sheet['E92'].value)) * 100 if sheet['E92'].value != 0 else 0
# 高价值降档占比
high_value_list1 = [cell.value for cell in sheet['J116':'O116'][0]]
high_value_list2 = [cell.value for cell in sheet['J92':'O92'][0]]
high_value_list3 = [(float(high_value_list1[i]) - float(high_value_list2[i])) / float(high_value_list2[i]) if float(high_value_list2[i]) != 0 else 0 for i in range(len(high_value_list2))]
high_value_regions = [cell.value for cell in sheet['J91':'O91'][0]]

# 找到high_value_list3中最大的两个值的下标
top_two_indices = sorted(range(len(high_value_list3)), key=lambda i: high_value_list3[i], reverse=True)[:2]

# 获取对应的high_value_regions值
top_two_regions = [high_value_regions[index] for index in top_two_indices]

# 业务自替代占比
self_substitution_ratio = (float(sheet['E183'].value) / float(sheet['E177'].value)) * 100 if float(sheet['E177'].value) != 0 else 0
self_substitution_list1 = [cell.value for cell in sheet['J183':'O183'][0]]
self_substitution_list2 = [cell.value for cell in sheet['J177':'O177'][0]]
self_substitution_list3 = [(float(self_substitution_list1[i]) - float(self_substitution_list2[i])) / float(self_substitution_list2[i]) if float(self_substitution_list2[i]) != 0 else 0 for i in range(len(self_substitution_list2))]
self_substitution_regions = high_value_regions
# 找到high_value_list3中最大的两个值的下标
top_two_indices2 = sorted(range(len(self_substitution_list3)), key=lambda i: self_substitution_list3[i], reverse=True)[:2]

# 获取对应的high_value_regions值
top_two_regions2 = [self_substitution_regions[index] for index in top_two_indices2]


# 异网抢夺占比
network_competition_ratio = (float(sheet['E187'].value) / float(sheet['E177'].value)) * 100 if float(sheet['E177'].value) != 0 else 0
network_competition_list1 = [cell.value for cell in sheet['J187':'O187'][0]]
network_competition_list2 = self_substitution_list2
network_competition_list3 = [(float(network_competition_list1[i]) - float(network_competition_list2[i])) / float(network_competition_list2[i]) if float(network_competition_list2[i]) != 0 else 0 for i in range(len(network_competition_list2))]
network_competition_regions = high_value_regions
# 找到high_value_list3中最大的两个值的下标
top_two_indices3 = sorted(range(len(network_competition_list3)), key=lambda i: network_competition_list3[i], reverse=True)[:2]

# 获取对应的high_value_regions值
top_two_regions3 = [network_competition_regions[index] for index in top_two_indices3]

# 集团降档占比
group_degraded_ratio = (float(sheet['E101'].value) / float(sheet['E92'].value)) * 100 if float(sheet['E92'].value) != 0 else 0
group_degraded_list1 = [cell.value for cell in sheet['J101':'O101'][0]]
group_degraded_list2 = high_value_list2
group_degraded_list3 = [(float(group_degraded_list1[i]) - float(group_degraded_list2[i])) / float(group_degraded_list2[i]) if float(group_degraded_list2[i]) != 0 else 0 for i in range(len(group_degraded_list2))]
group_degraded_regions = high_value_regions
# 找到high_value_list3中最大的两个值的下标
top_two_indices4 = sorted(range(len(group_degraded_list3)), key=lambda i: group_degraded_list3[i], reverse=True)[:2]

# 获取对应的high_value_regions值
top_two_regions4 = [group_degraded_regions[index] for index in top_two_indices4]
'''