from openpyxl import load_workbook

# 文件路径
file_path = r"F:\\IPA_project\\网格周体验单生成流程\\data\\周体检报告（县格）-20240706修正.xlsx"

# 加载工作簿
workbook = load_workbook(filename=file_path)

# 指定需要处理的工作表名称
sheets_to_process = ['数据源【当周】', '数据源【上周】', '小福包【当周】', '小福包【上周】']

# 遍历工作表
for sheet_name in sheets_to_process:
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        # 假设列头在第3行
        header_row = 3

        # 保存列头
        headers = [cell.value for cell in worksheet[header_row]]

        # 清空数据，但保留列头
        for row in worksheet.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.value = None

        # 确保列头仍然存在
        for i, header in enumerate(headers, start=1):
            worksheet.cell(row=header_row, column=i).value = header

# 保存修改后的工作簿
workbook.save('处理后的_周体检报告（县格）-20240706修正.xlsx')

# 关闭工作簿
workbook.close()