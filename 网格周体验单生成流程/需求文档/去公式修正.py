import win32com.client as win32
from openpyxl import load_workbook
from copy import copy
def just_open(filename):
    xl_app = win32.Dispatch('Ket.Application')    # 如果是WPS则改成('Ket.Application')
    xl_app.Visible = False
    xl_book = xl_app.Workbooks.Open(filename, UpdateLinks=False, ReadOnly=False)
    xl_book.Save()
    xl_book.Close()
path_to_your_file = r"F:\\IPA_project\\网格周体验单生成流程\\需求文档\\周体检报告（县格）【无公式版】-20240722修正.xlsx"
just_open(path_to_your_file)

# 加载工作簿
wb = load_workbook(filename=path_to_your_file, data_only=False)  # 不设置data_only=True，以便读取公式

# 创建一个新工作簿来保存修改后的工作表
new_wb = load_workbook(filename=path_to_your_file, data_only=True)  # 设置data_only=True，以便只读取值

# 遍历工作簿中的每个工作表
for sheetname in wb.sheetnames:
    # 获取原工作表和新工作簿中的工作表
    original_ws = wb[sheetname]
    new_ws = new_wb[sheetname]

    # 检查工作表名称，如果是“查询【市县】”或“查询【网格】”，则不做处理
    if sheetname in ['查询【市县】', '查询【网格】']:

        # 遍历原工作表中的每个单元格
        for row in original_ws.iter_rows(min_row=1, max_col=original_ws.max_column, max_row=original_ws.max_row):
            for cell in row:
                # 获取新工作表中的对应单元格
                new_cell = new_ws[cell.coordinate]

                # 如果单元格包含公式
                if cell.data_type == 'f':
                    # 将计算结果复制到新工作表的单元格
                    new_cell.value = cell.value
                    # 复制单元格样式
                    new_cell._style = copy(cell._style)
    # 否则，如果工作表是“查询【市县】”或“查询【网格】”，保持原样

# 保存新工作簿
new_wb.save('周体检报告（县格）【无公式版】-20240722修正-1')
