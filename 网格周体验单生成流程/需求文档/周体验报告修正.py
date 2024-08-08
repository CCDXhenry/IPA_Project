import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# 获取当前日期
date = datetime.today()
date_str = date.strftime("%Y%m%d")

# 源文件和目标文件路径
source_file = r"网格调度测试-数据源版（周）_20240721_17_00_00.xlsx"
target_file = r"周体检报告（县格）-20240720修正.xlsx"

# 打开目标文件工作簿
target_wb = load_workbook(target_file)

# 要处理的工作表列表
sheets_to_process = ['数据源【当周】', '数据源【上周】', '小福包【当周】', '小福包【上周】']

# 遍历每个工作表
for sheet_name in sheets_to_process:
    if sheet_name in target_wb.sheetnames:
        # 读取目标工作表
        target_ws = target_wb[sheet_name]

        # 读取源文件前三行数据
        source_header_df = pd.read_excel(source_file, sheet_name=sheet_name, nrows=3, header=None)

        # 读取目标文件前三行数据
        target_header_df = pd.read_excel(target_file, sheet_name=sheet_name, nrows=3, header=None)

        # 清除数据，但保留公式
        for row in target_ws.iter_rows(min_row=4):  # 从第二行开始，假设第一行为标题
            for cell in row:
                if cell.data_type == 'f':  # 如果单元格包含公式
                    pass  # 不做任何操作，保留公式
                else:
                    cell.value = None  # 清除非公式内容

        # 获取源数据框和目标数据框的列数
        num_source_cols = source_header_df.shape[1]
        num_target_cols = target_header_df.shape[1]

        # 初始化列匹配字典
        col_match = {}

        # 寻找匹配的列
        for source_col_idx in range(num_source_cols):
            match_found = False
            for target_col_idx in range(num_target_cols):
                # 检查列数据是否相等
                if source_header_df.iloc[:, source_col_idx].equals(target_header_df.iloc[:, target_col_idx]):
                    col_match[source_col_idx] = target_col_idx
                    match_found = True
                    break
            if not match_found:
                print(f"未找到匹配列: {sheet_name} 源列 {source_col_idx}")

        # 现在col_match字典包含了源列索引到目标列索引的映射
        for source_col_idx, target_col_idx in col_match.items():
            print(f"匹配成功: {sheet_name} 源列 {source_col_idx} 目标列 {target_col_idx}")

            # 读取源文件剩余数据的这一列
            source_col_df = pd.read_excel(source_file, sheet_name=sheet_name, usecols=[source_col_idx], skiprows=3,
                                          header=None)
            # 在读取源文件剩余数据的这一列之后，检查DataFrame的内容
            print(f"读取的源列数据: {source_col_df}")

            # 将数据粘贴到目标文件
            for index, row in source_col_df.iterrows():
                #print(type(row))  # 应输出 <class 'pandas.core.series.Series'>
                #print(row)  # 打印整个row，确认其内容
                value = row.iloc[0]
                #print(value)
                target_ws.cell(row=index + 4, column=target_col_idx + 1, value=value)


# 保存目标工作簿
target_wb.save(f'周体检报告（县格）【无公式版】-{date_str}修正.xlsx')

# 关闭工作簿
target_wb.close()