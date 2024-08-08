from openpyxl import load_workbook

# 读取source_file中的“通路对应表”
path_to_source_file = '网格调度测试-数据源版（周）_20240708_15_00_00.xlsx'
wb_source = load_workbook(filename=path_to_source_file, data_only=True)
source_sheet = wb_source['通路对应表']

# 读取目标工作表
path_to_target_file = '周体检报告（县格）-20240706修正数据黏贴.xlsx'
path_to_target_file_end = '周体检报告（县格）-20240706修正网格编码匹配.xlsx'
wb_target = load_workbook(filename=path_to_target_file, data_only=True)

# 要处理的工作表列表
sheets_to_process = ['数据源【当周】', '数据源【上周】', '小福包【当周】', '小福包【上周】']

# 遍历每个工作表
for sheet_name in sheets_to_process:
    if sheet_name in wb_target.sheetnames:
        target_sheet = wb_target[sheet_name]

        # 找到网格编码所在的列，这里假设在C列（索引为2）
        grid_code_column = 12
        target_row_index = 4
        # 从第四行开始遍历
        for row_idx, row in enumerate(target_sheet.iter_rows(min_row=4, values_only=True), start=4):
            grid_code = row[grid_code_column]

            if grid_code is not None and grid_code != 0:
                # 在“通路对应表”中查找匹配的CHANNEL_ID
                for source_row in source_sheet.iter_rows(values_only=True):
                    if str(source_row[1]) == str(grid_code):  # 假设CHANNEL_ID在A列（索引为0）
                        chnl_level_1 = source_row[8]  # CHNL_LEVEL_1在B列（索引为1）
                        chnl_level_2 = source_row[9]  # CHNL_LEVEL_2在C列（索引为2）
                        short_name = source_row[0]  # 社渠简称为D列（索引为3）
                        print(chnl_level_1)
                        print(chnl_level_2)
                        print(short_name)
                        # 更新目标工作表
                        #target_row_index = row[0]  # 这里假设你需要找到目标行的索引
                        target_sheet.cell(row=target_row_index, column=1).value = chnl_level_1
                        target_sheet.cell(row=target_row_index, column=2).value = chnl_level_2
                        target_sheet.cell(row=target_row_index, column=3).value = short_name
                        target_row_index += 1
                        break

# 保存更新后的目标工作表
wb_target.save(path_to_target_file_end)
