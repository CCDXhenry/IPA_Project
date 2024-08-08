from openpyxl import load_workbook
from datetime import datetime

# 获取当前日期
date = datetime.today()
date_str = date.strftime("%Y%m%d")

# 加载工作簿
source_wb = load_workbook('网格调度测试-数据源版（周）_20240708_15_00_00.xlsx')
target_wb = load_workbook('周体检报告（县格）-20240706修正数据黏贴.xlsx')

# 选择工作表
target_ws = target_wb['通路对应表']
source_ws = source_wb['通路对应表']

# 获取目标表的列标题
target_header = [cell.value for cell in target_ws[2]]
print(target_header)
column_mapping = {
    'CHANNEL_NAME': 'BOSS渠道名称',
    'CHANNEL_ID': 'BOSS渠道编码',
    'GRID_NAME': '归属网格',
    'CHNL_LEVEL_1': '渠道分层1',
    'CHNL_LEVEL_2': '渠道分层2',
    '社渠简称': '社会渠道商抓取'
}
# 创建一个字典来存储目标表的列索引
target_col_indices = {col: idx + 1 for idx, col in enumerate(target_header) if col in column_mapping.values()}

# 创建一个字典来存储源表的列索引
source_col_indices = {col: idx + 1 for idx, col in enumerate(next(source_ws.iter_rows(values_only=True))) if col in column_mapping.keys()}

# 创建一个最终的映射，将目标列索引映射到源列索引
column_map_final = {target_col_indices[val]: source_col_indices[key] for key, val in column_mapping.items() if key in source_col_indices and val in target_col_indices}

# 清除目标表中除了前两行以外的所有数据
rows_to_delete = list(target_ws.rows)[2:]
for row in rows_to_delete:
    target_ws.delete_rows(idx=row[0].row)

# 开始从第二行开始添加数据
row_idx = 3
for src_row in source_ws.iter_rows(min_row=2, values_only=True):
    for target_col, source_col in column_map_final.items():
        target_ws.cell(row=row_idx, column=target_col).value = src_row[source_col - 1]  # 减1是因为列表索引从0开始
    row_idx += 1

# 保存目标工作簿
target_wb.save(f'周体检报告（县格）【无公式版】-{date_str}修正.xlsx')