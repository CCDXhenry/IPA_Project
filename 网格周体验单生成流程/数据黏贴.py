import pandas as pd
from openpyxl import load_workbook

# 源文件路径
source_file = '网格调度测试-数据源版（周）_20240708_15_00_00.xlsx'
# 目标文件路径
target_file = '处理后的_周体检报告（县格）-20240706修正.xlsx'
target_file_end = '周体检报告（县格）-20240706修正数据黏贴.xlsx'
# 读取源文件中的数据
source_sheets = ['数据源【当周】', '数据源【上周】', '小福包【当周】', '小福包【上周】']
source_data = {sheet: pd.read_excel(source_file, sheet_name=sheet) for sheet in source_sheets}


# 读取源文件中的“通路对应表”
mapping_df = pd.read_excel(source_file, sheet_name='通路对应表')

# 创建映射字典
mapping_dict = mapping_df.set_index('CHANNEL_ID')[['CHNL_LEVEL_1', 'CHNL_LEVEL_2', '社渠简称']].to_dict(orient='index')

# 加载目标文件
target_wb = load_workbook(target_file)

# 数据粘贴
for sheet_name, data in source_data.items():
    if sheet_name in target_wb.sheetnames:
        target_ws = target_wb[sheet_name]

        # 数据开始的行列号
        start_row = 2
        start_col = 4
        data = data.iloc[2:]
        # 写入数据
        for idx, row in data.iterrows():
            for col_idx, value in enumerate(row, start=1):
                target_ws.cell(row=start_row + idx, column=start_col + col_idx - 1, value=value)

for sheet_name in source_sheets:
    if sheet_name in target_wb.sheetnames:
        target_ws = target_wb[sheet_name]

        # 数据开始的行列号
        start_row = 4
        start_col = 4

        # 遍历行
        for row in target_ws.iter_rows(min_row=start_row, min_col=start_col, max_col=start_col + 2, values_only=True):
            channel_id = row[4 - start_col]  # 第9列的值，注意列偏移

            # 根据channel_id查找映射值
            if channel_id in mapping_dict:
                level_1, level_2, short_name = mapping_dict[channel_id]['CHNL_LEVEL_1'], mapping_dict[channel_id][
                    'CHNL_LEVEL_2'], mapping_dict[channel_id]['社渠简称']

                # 写入数据
                target_ws.cell(row=row[0], column=start_col - 3).value = level_1
                target_ws.cell(row=row[0], column=start_col - 2).value = level_2
                target_ws.cell(row=row[0], column=start_col - 1).value = short_name

# 保存文件
target_wb.save(target_file_end)

# 关闭工作簿
target_wb.close()